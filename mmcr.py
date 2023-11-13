from selenium import webdriver
from requestium import Session
from time import sleep
import eLead
import pandas as pd
import json
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment

mmcr = "https://mmcr-amap.i.mercedes-benz.com"


def loginDDA(s,user,password):
    """Helper function for logging into MMCR."""

    #get mmcr directly, avoid netstar as mmcr directly requires no 2FA
    s.driver.get(mmcr)

    s.driver.ensure_element_by_id("userid").send_keys(user)
    s.driver.ensure_element_by_id("next-btn").click()
    s.driver.ensure_element_by_id("password").send_keys(password)
    s.driver.ensure_element_by_id("loginSubmitButton").click()
    sleep(6)
    s.copy_user_agent_from_driver()
    s.transfer_driver_cookies_to_session()
    s.driver.minimize_window()

    return s


def vinSearch(s,vin):
    status = '?'
    expiry = 'N/A'

    r = s.get("https://mmcr-amap.i.mercedes-benz.com/api/vehicles?finOrVin="+vin+"&country=US&language=en-US")
    if r.status_code!=200:
        status = "Bad Response"
        return status,expiry
    else:
        respJSON = json.loads(r.content)
        try:
            accountID = respJSON['accountRoles'][0]
        except:
            status = "Not Paired"
            return status,expiry
        accountID = accountID['accountId']

        r = s.get("https://mmcr-amap.i.mercedes-benz.com/api/vehicles/services?finOrVin="+vin+"&language=en-US&accountId="+accountID)
        if r.status_code == 200:
            servicesJSON = json.loads(r.content)
            statusList = servicesJSON['services']
            for service in statusList:
                if service['description'] == "Remote Engine Start":
                    status = service["activationStatus"]
                    if status == "ACTIVE":
                        status = "Active"
                        expiry = service["license"]
                        expiry = expiry["end"]
                        expiry = expiry[0:10]
                    break
            if status == "INACTIVE":
                status = "Inactive"
    return status,expiry


def daysBetween(expiry):
    """Function takes in a list of expiry dates and determines how many days between now and expiry,
      then adds how many days until expiry next to expiry date in the list."""
    
    today = date.today()

    for i,expiryDate in enumerate(expiry):
        if str(expiryDate)=="N/A":
            continue
        else:
            split = str(expiryDate)[0:10].split('-')
            year = int(split[0])
            month = int(split[1])
            day = int(split[2])
            d = date(year,month,day)
            dayNum = (d - today).days
            expiry[i] = str(expiry[i])[0:10] +' (' + str(dayNum) + ' days)'
    return expiry


def organize(indexList, path='out.xlsx'):

    wb = load_workbook(path)
    sheet = wb.active
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    b = sheet['A1':'G'+str(sheet.max_row)]

    for cell in b:
        for x in cell:
            x.border=thin_border
            x.alignment = Alignment(horizontal='center')

    for offset,i in enumerate(indexList):
        sheet.insert_rows(i + offset)
        
    wb.save(path)


def main():
    email = input("Please input your email (ending in @mbnaples.com):")
    username = input("Please input your netstar login (starting with D7):")
    ePassword = input("Please input your eLeads password:")
    mPassword = input("Please input your Netstar password:")

    eLeadDriver = webdriver.Chrome()
    df = eLead.getSchedule(eLeadDriver,email,ePassword)
    eLeadDriver.quit()
    
    driver = webdriver.Chrome()
    vinList = df['VIN'].to_list()
    statusList = []
    expiryList = []

    s = Session(webdriver_path='./chromedriver',driver=driver)
    s = loginDDA(s,username,mPassword)
    print(" ")
    for v in vinList:
        statusTemp,expiryTemp = vinSearch(s,v)
        statusList.append(statusTemp)
        if statusTemp == "Bad Response" or statusTemp == '?':
            print(v + ": " + statusTemp)
        elif statusTemp == "Not Paired":
            print(v + ": " + statusTemp)
        else:
            print(v + ": " + statusTemp)

        expiryList.append(expiryTemp)

    dropList = []
    for i,expiry in enumerate(expiryList):
        if expiry!='N/A':
            split = expiry.split('-')
            split[2] = split[2][:4]
            year = int(split[0])
            month = int(split[1])
            day = int(split[2])
            dayy = date(year,month,day)
            today = dayy.today()
            if (dayy-today).days>90:
                dropList.append(i)

    df['Status'] = statusList
    df['Expiry'] = daysBetween(expiryList)

    df = df.drop(index=dropList)
    df = df.drop(['VIN'],axis=1)
    df.sort_values(by='Service Rep', inplace=True, kind='stable')
    idxList = []
    repList = df["Service Rep"].to_list()
    prevRep = ""

    for idx,currentRep in enumerate(repList):
        if currentRep != prevRep:
            prevRep = currentRep
            if idx != 0:
                idxList.append(idx+2)

    print(" ")
    try:
        with pd.ExcelWriter("out.xlsx",mode="w") as writer:
            df.to_excel(excel_writer = writer,index=False)
        print("Successfully Saved To: out.xlsx")
    except:
        print("Failed to write to excel file. Please ensure out.xlsx is not open whilst running")
        sleep(5)

    organize(idxList)